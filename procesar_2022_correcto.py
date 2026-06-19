#!/usr/bin/env python3
from openpyxl import load_workbook
from datetime import datetime
from collections import defaultdict
import re

# Cargar el archivo
wb = load_workbook('agendamientos.xlsx')
ws = wb['2022']

# Diccionario para almacenar datos por mes
months_data = defaultdict(lambda: {'consultas': [], 'consultantes_set': set()})

# Primero, encontrar todas las fechas en la hoja (Fila 3 tiene las fechas)
fecha_row = 3
fechas_por_col = {}

for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=fecha_row, column=col).value
    if isinstance(cell, datetime):
        fechas_por_col[col] = cell

print(f"Encontradas {len(fechas_por_col)} fechas\n")

# Procesardata de consultas (filas 4 en adelante, omitiendo encabezados)
for row_idx in range(4, ws.max_row + 1):
    horario_cell = ws.cell(row=row_idx, column=1).value

    # Si es un horario válido (termina con "hs")
    if horario_cell and isinstance(horario_cell, str) and 'hs' in horario_cell:
        # Procesar cada columna en esta fila
        for col in range(2, ws.max_column + 1):
            nombre_cell = ws.cell(row=row_idx, column=col).value

            # Si hay un nombre válido (no "-", no vacío)
            if nombre_cell and isinstance(nombre_cell, str):
                nombre = nombre_cell.strip()
                if nombre and nombre != '-':
                    # Limpiar el nombre
                    nombre = nombre.replace(' VIR', '').replace(' virtual', '').replace(' videollamada', '')
                    nombre = re.sub(r'\s\d{1,2}:\d{2}', '', nombre).strip()

                    # Obtener la fecha de esta columna
                    if col in fechas_por_col:
                        fecha = fechas_por_col[col]
                        mes = fecha.month

                        # Agregar a la base de datos
                        months_data[mes]['consultas'].append(nombre)
                        months_data[mes]['consultantes_set'].add(nombre.lower())

# Mostrar resultados
print("="*70)
print("ANÁLISIS 2022 - CONTEO CORRECTO DESDE EXCEL")
print("="*70)

meses_nombres = {
    8: 'Agosto',
    9: 'Septiembre',
    10: 'Octubre',
    11: 'Noviembre',
    12: 'Diciembre'
}

for mes in sorted(months_data.keys()):
    data = months_data[mes]
    total_consultas = len(data['consultas'])
    consultantes_distintos = len(data['consultantes_set'])

    mes_nombre = meses_nombres.get(mes, f'Mes {mes}')
    print(f"\n{mes_nombre.upper()} ({mes}/2022)")
    print(f"  Total consultas (celdas con nombres): {total_consultas}")
    print(f"  Consultantes distintos: {consultantes_distintos}")

print("\n" + "="*70)
print("RESUMEN PARA AIRTABLE")
print("="*70)
print("\nMes | Año | Consultantes Distintos | Total Consultas | Canceladas | Tasa %")
print("-" * 70)

for mes in sorted(months_data.keys()):
    data = months_data[mes]
    mes_nombre = meses_nombres.get(mes, f'Mes {mes}')
    consultantes = len(data['consultantes_set'])
    consultas = len(data['consultas'])
    print(f"{mes_nombre} | 2022 | {consultantes} | {consultas} | 0 | 0%")
