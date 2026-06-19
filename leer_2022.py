#!/usr/bin/env python3
from openpyxl import load_workbook

# Cargar el archivo
wb = load_workbook('agendamientos.xlsx')

# Listar hojas disponibles
print("Hojas disponibles:")
for sheet_name in wb.sheetnames:
    print(f"  - {sheet_name}")

# Verificar si existe la hoja "2022"
if '2022' in wb.sheetnames:
    ws = wb['2022']
    print(f"\nHoja 2022 encontrada")
    print(f"Dimensiones: {ws.dimensions}")

    # Ver las primeras 20 filas
    print("\nPrimeras 20 filas (primeras 8 columnas):")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        print(f"Fila {i}: {row[:8]}")
else:
    print(f"\nHoja 2022 NO encontrada")
    print(f"Hojas disponibles: {wb.sheetnames}")
