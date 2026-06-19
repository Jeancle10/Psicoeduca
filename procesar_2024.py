#!/usr/bin/env python3
from openpyxl import load_workbook
from datetime import datetime
from collections import defaultdict
import re

# Cargar el archivo
wb = load_workbook('agendamientos.xlsx')

# Verificar hoja 2024
if '2024' not in wb.sheetnames:
    print("ERROR: Hoja 2024 NO encontrada")
    print(f"Hojas disponibles: {wb.sheetnames}")
    exit(1)

ws = wb['2024']
print(f"Procesando hoja 2024 (Dimensiones: {ws.dimensions})\n")

# Diccionario para almacenar datos por mes
months_data = defaultdict(lambda: {'consultas': [], 'consultantes_set': set()})

# PASO 1: Encontrar todas las filas que contienen fechas
fecha_rows = []

for row_idx in range(1, ws.max_row + 1):
    fecha_count = 0
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col).value
        if isinstance(cell, datetime):
            fecha_count += 1

    if fecha_count >= 5:
        fecha_rows.append(row_idx)
        print(f"Bloque de fechas encontrado en Fila {row_idx}")

print(f"\nTotal de bloques de meses encontrados: {len(fecha_rows)}\n")

# PASO 2: Procesar cada bloque de fechas
for block_idx, fecha_row in enumerate(fecha_rows):
    print(f"Procesando bloque {block_idx + 1} (Fila {fecha_row})...")

    # Extraer fechas
    fechas_por_col = {}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=fecha_row, column=col).value
        if isinstance(cell, datetime):
            fechas_por_col[col] = cell

    if fechas_por_col:
        first_date = list(fechas_por_col.values())[0]
        print(f"  Mes: {first_date.strftime('%B %Y')}")

    # Procesar filas de datos
    next_fecha_row = fecha_rows[block_idx + 1] if block_idx + 1 < len(fecha_rows) else ws.max_row + 1
    data_rows = range(fecha_row + 1, next_fecha_row)

    for row_idx in data_rows:
        # Procesar todos los nombres en la fila, sin requerir horario explícito
        for col in range(2, ws.max_column + 1):
            nombre_cell = ws.cell(row=row_idx, column=col).value

            if nombre_cell and isinstance(nombre_cell, str):
                nombre = nombre_cell.strip()
                if nombre and nombre != '-':
                    # Limpiar nombre
                    nombre = nombre.replace(' VIR', '').replace(' virtual', '').replace(' videollamada', '')
                    nombre = re.sub(r'\s\d{1,2}:\d{2}', '', nombre).strip()
                    nombre = nombre.replace(' P/', '').strip()

                    if col in fechas_por_col:
                        fecha = fechas_por_col[col]
                        mes = fecha.month

                        months_data[mes]['consultas'].append(nombre)
                        months_data[mes]['consultantes_set'].add(nombre.lower())

# Mostrar resultados
print("\n" + "="*70)
print("ANÁLISIS 2024 - DESDE EXCEL")
print("="*70)

meses_nombres = {
    1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo',
    6: 'Junio', 7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre',
    11: 'Noviembre', 12: 'Diciembre'
}

for mes in range(1, 13):
    if mes in months_data:
        data = months_data[mes]
        total_consultas = len(data['consultas'])
        consultantes_distintos = len(data['consultantes_set'])
    else:
        total_consultas = 0
        consultantes_distintos = 0

    mes_nombre = meses_nombres.get(mes, f'Mes {mes}')
    print(f"\n{mes_nombre.upper()} ({mes}/2024)")
    print(f"  Total consultas: {total_consultas}")
    print(f"  Consultantes distintos: {consultantes_distintos}")

print("\n" + "="*70)
print("PARA AIRTABLE")
print("="*70)
print("\nMes | Consultantes | Consultas")
print("-" * 50)

for mes in range(1, 13):
    if mes in months_data:
        data = months_data[mes]
        mes_nombre = meses_nombres.get(mes, f'Mes {mes}')
        print(f"{mes_nombre} | {len(data['consultantes_set'])} | {len(data['consultas'])}")
    else:
        mes_nombre = meses_nombres.get(mes, f'Mes {mes}')
        print(f"{mes_nombre} | 0 | 0")
