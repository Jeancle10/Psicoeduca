#!/usr/bin/env python3
from openpyxl import load_workbook
from datetime import datetime
from collections import defaultdict

# Cargar el archivo
wb = load_workbook('agendamientos.xlsx')
ws = wb['2022']

# Función para extraer mes de una fecha
def get_month(date_obj):
    if isinstance(date_obj, datetime):
        return date_obj.month
    return None

# Diccionario para almacenar datos por mes
months_data = defaultdict(lambda: {'consultas': [], 'consultantes_set': set()})

# Procesar todas las filas
print("Procesando hoja 2022...\n")

# Iterar sobre todas las filas
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    # Saltar filas vacías
    if not any(row):
        continue

    # La primera columna es el horario
    horario = row[0]

    # Si comienza con un número de hora, procesar la fila
    if horario and isinstance(horario, str) and 'hs' in horario:
        # Procesar cada celda (excepto la primera que es el horario)
        for col_idx, cell in enumerate(row[1:], start=1):
            # Obtener la fecha (normalmente está 2 filas arriba)
            date_row = ws.cell(row=ws._current_row - 3, column=col_idx + 1).value if hasattr(ws, '_current_row') else None

            # Buscar la fecha más cercana en las filas anteriores
            fecha = None
            for search_row in range(ws.cell(row=ws._current_row - 1, column=col_idx + 1).row if hasattr(ws, '_current_row') else 3, 0, -1):
                cell_val = ws.cell(row=search_row, column=col_idx + 1).value
                if isinstance(cell_val, datetime):
                    fecha = cell_val
                    break

            # Si la celda tiene un nombre (no está vacía ni es "-")
            if cell and isinstance(cell, str) and cell.strip() and cell.strip() != '-':
                nombre = cell.strip()

                # Limpiar el nombre (remover "VIR", "virtual", horas, etc.)
                nombre = nombre.replace(' VIR', '').replace(' virtual', '').replace(' videollamada', '')
                # Remover números entre paréntesis o horas
                import re
                nombre = re.sub(r'\s\d{1,2}:\d{2}', '', nombre).strip()

                if nombre and fecha:
                    mes = get_month(fecha)
                    if mes:
                        months_data[mes]['consultas'].append(nombre)
                        months_data[mes]['consultantes_set'].add(nombre.lower())

# Mostrar resultados
print("="*70)
print("ANÁLISIS 2022 - CONTEO DIRECTO DE HOJA EXCEL")
print("="*70)

meses_nombres = {
    8: 'Agosto',
    9: 'Septiembre',
    10: 'Octubre',
    11: 'Noviembre',
    12: 'Diciembre'
}

for mes in sorted(months_data.keys()):
    data = months_data[mes]
    total_consultas = len(data['consultas'])
    consultantes_distintos = len(data['consultantes_set'])

    mes_nombre = meses_nombres.get(mes, f'Mes {mes}')
    print(f"\n{mes_nombre.upper()} ({mes}/2022)")
    print(f"  Total consultas: {total_consultas}")
    print(f"  Consultantes distintos: {consultantes_distintos}")
    print(f"  Primeros 10 consultantes: {sorted(data['consultantes_set'])[:10]}")

print("\n" + "="*70)
print("RESUMEN TOTALES")
print("="*70)

for mes in sorted(months_data.keys()):
    data = months_data[mes]
    mes_nombre = meses_nombres.get(mes, f'Mes {mes}')
    print(f"{mes_nombre}: {len(data['consultantes_set'])} consultantes | {len(data['consultas'])} consultas")
