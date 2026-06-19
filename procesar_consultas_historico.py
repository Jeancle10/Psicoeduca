#!/usr/bin/env python3
"""
Extrae CADA CONSULTA INDIVIDUAL (nombre, fecha exacta, modalidad) de todas las hojas
y genera datos para cargar en Airtable tabla "Consultas Histórico"
"""
from openpyxl import load_workbook
from datetime import datetime
import re
import json

wb = load_workbook('agendamientos.xlsx')

# Diccionario para almacenar todas las consultas
todas_consultas = []

for año_str in ['2022', '2023', '2024', '2025', '2026']:
    if año_str not in wb.sheetnames:
        print(f"⚠️ Hoja {año_str} no encontrada, saltando...")
        continue

    print(f"\n{'='*70}")
    print(f"PROCESANDO {año_str}")
    print(f"{'='*70}")

    ws = wb[año_str]
    año = int(año_str)

    # PASO 1: Encontrar bloques de fechas (encabezados de meses)
    fecha_rows = []
    for row_idx in range(1, ws.max_row + 1):
        fecha_count = 0
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col).value
            if isinstance(cell, datetime):
                fecha_count += 1
        if fecha_count >= 5:
            fecha_rows.append(row_idx)

    print(f"Bloques de meses encontrados: {len(fecha_rows)}")

    # PASO 2: Procesar cada bloque
    for block_idx, fecha_row in enumerate(fecha_rows):
        # Extraer fechas del encabezado
        fechas_por_col = {}
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=fecha_row, column=col).value
            if isinstance(cell, datetime):
                fechas_por_col[col] = cell

        if not fechas_por_col:
            continue

        first_date = list(fechas_por_col.values())[0]
        mes = first_date.month
        mes_nombre = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                      'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][mes]

        print(f"\n  {mes_nombre} {año}:")

        # Procesar filas de datos
        next_fecha_row = fecha_rows[block_idx + 1] if block_idx + 1 < len(fecha_rows) else ws.max_row + 1
        data_rows = range(fecha_row + 1, next_fecha_row)

        contador = 0
        for row_idx in data_rows:
            # Procesar cada nombre en cada columna
            for col in range(2, ws.max_column + 1):
                nombre_cell = ws.cell(row=row_idx, column=col).value

                if nombre_cell and isinstance(nombre_cell, str):
                    nombre = nombre_cell.strip()
                    if nombre and nombre != '-':
                        # Detectar modalidad
                        es_virtual = 'VIR' in nombre.upper()

                        # Limpiar nombre
                        nombre_limpio = nombre.replace(' VIR', '').replace(' virtual', '').replace(' videollamada', '')
                        nombre_limpio = re.sub(r'\s\d{1,2}:\d{2}', '', nombre_limpio).strip()
                        nombre_limpio = nombre_limpio.replace(' P/', '').strip()

                        # Si tiene fecha en esa columna
                        if col in fechas_por_col:
                            fecha = fechas_por_col[col]

                            # Crear registro de consulta
                            consulta = {
                                'nombre': nombre_limpio,
                                'fecha': fecha.strftime('%Y-%m-%d'),
                                'año': año,
                                'mes': mes,
                                'mes_nombre': mes_nombre,
                                'modalidad': 'Virtual' if es_virtual else 'Presencial'
                            }
                            todas_consultas.append(consulta)
                            contador += 1

        print(f"    {contador} consultas procesadas")

print(f"\n{'='*70}")
print(f"RESUMEN FINAL")
print(f"{'='*70}")
print(f"Total consultas procesadas: {len(todas_consultas)}")
print(f"Años: 2022-2026")

# Agrupar por año
por_año = {}
for consulta in todas_consultas:
    año = consulta['año']
    if año not in por_año:
        por_año[año] = 0
    por_año[año] += 1

for año in sorted(por_año.keys()):
    print(f"  {año}: {por_año[año]} consultas")

# Guardar en JSON para verificación
with open('consultas_historico.json', 'w', encoding='utf-8') as f:
    json.dump(todas_consultas, f, ensure_ascii=False, indent=2)

print(f"\n✅ Datos guardados en: consultas_historico.json")

# Mostrar primeras 5 y últimas 5 consultas
print(f"\n{'='*70}")
print(f"MUESTRA DE DATOS (primeras 5)")
print(f"{'='*70}")
for consulta in todas_consultas[:5]:
    print(f"{consulta['nombre']:30} | {consulta['fecha']} | {consulta['modalidad']:10}")

print(f"\n{'='*70}")
print(f"MUESTRA DE DATOS (últimas 5)")
print(f"{'='*70}")
for consulta in todas_consultas[-5:]:
    print(f"{consulta['nombre']:30} | {consulta['fecha']} | {consulta['modalidad']:10}")
