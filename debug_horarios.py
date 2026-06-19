from openpyxl import load_workbook
from datetime import time

wb = load_workbook('agendamientos.xlsx')
ws = wb['2024']

# Revisar toda la columna A en el rango de abril (44-58)
print('=== Todos los horarios en abril (filas 44-58) ===')
for row in range(44, 59):
    h = ws.cell(row=row, column=1).value
    print(f'Fila {row}: {type(h).__name__:15} = {h}')

print('\n=== Verificar nombres en fila 45 ===')
for col in range(1, 15):
    cell = ws.cell(row=45, column=col).value
    print(f'Col {col:2}: {cell}')
