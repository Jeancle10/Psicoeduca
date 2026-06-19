#!/usr/bin/env python3
from openpyxl import load_workbook
from datetime import datetime
from collections import defaultdict
import re

# Cargar el archivo
wb = load_workbook('agendamientos.xlsx')
ws = wb['2022']

# Diccionario para almacenar datos por mes
months_data = defaultdict(lambda: {'consultas': [], 'consultantes_set': set()})

# PASO 1: Encontrar todas las filas que contienen fechas (estos son los encabezados de cada mes)
fecha_rows = []

for row_idx in range(1, ws.max_row + 1):
    # Contar cuántas fechas hay en esta fila
    fecha_count = 0
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col).value
        if isinstance(cell, datetime):
            fecha_count += 1

    # Si hay al menos 5 fechas en esta fila, es una fila de encabezado
    if fecha_count >= 5:
        fecha_rows.append(row_idx)
        print(f"Encontrada fila de fechas en: Fila {row_idx}")

print(f"\nTotal de bloques de meses encontrados: {len(fecha_rows)}\n")

# PASO 2: Para cada bloque de fechas, procesar los datos
for block_idx, fecha_row in enumerate(fecha_rows):
    print(f"Procesando bloque {block_idx + 1} (Fila {fecha_row})...")

    # Extraer fechas de esta fila
    fechas_por_col = {}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=fecha_row, column=col).value
        if isinstance(cell, datetime):
            fechas_por_col[col] = cell

    print(f"  Fechas encontradas: {len(fechas_por_col)}")
    if fechas_por_col:
        first_date = list(fechas_por_col.values())[0]
        print(f"  Mes: {first_date.strftime('%B %Y')}")

    # Procesar las filas siguientes (hasta encontrar otra fila de fechas o fin)
    next_fecha_row = fecha_rows[block_idx + 1] if block_idx + 1 < len(fecha_rows) else ws.max_row + 1
    data_rows = range(fecha_row + 1, next_fecha_row)

    # Procesar cada fila de datos
    for row_idx in data_rows:
        horario_cell = ws.cell(row=row_idx, column=1).value

        # Si es un horario válido
        if horario_cell and isinstance(horario_cell, str) and 'hs' in horario_cell:
            # Procesar cada columna
            for col in range(2, ws.max_column + 1):
                nombre_cell = ws.cell(row=row_idx, column=col).value

                # Si hay un nombre válido
                if nombre_cell and isinstance(nombre_cell, str):
                    nombre = nombre_cell.strip()
                    if nombre and nombre != '-':
                        # Limpiar el nombre
                        nombre = nombre.replace(' VIR', '').replace(' virtual', '').replace(' videollamada', '')
                        nombre = re.sub(r'\s\d{1,2}:\d{2}', '', nombre).strip()
                        nombre = nombre.replace(' P/', '').strip()

                        # Obtener la fecha
                        if col in fechas_por_col:
                            fecha = fechas_por_col[col]
                            mes = fecha.month

                            # Agregar
                            months_data[mes]['consultas'].append(nombre)
                            months_data[mes]['consultantes_set'].add(nombre.lower())

# Mostrar resultados
print("\n" + "="*70)
print("ANÁLISIS 2022 - DESDE EXCEL (FINAL)")
print("="*70)

meses_nombres = {
    8: 'Agosto',
    9: 'Septiembre',
    10: 'Octubre',
    11: 'Noviembre',
    12: 'Diciembre'
}

for mes in sorted(months_data.keys()):
    data = months_data[mes]
    total_consultas = len(data['consultas'])
    consultantes_distintos = len(data['consultantes_set'])

    mes_nombre = meses_nombres.get(mes, f'Mes {mes}')
    print(f"\n{mes_nombre.upper()} ({mes}/2022)")
    print(f"  Total consultas: {total_consultas}")
    print(f"  Consultantes distintos: {consultantes_distintos}")

print("\n" + "="*70)
print("PARA AIRTABLE")
print("="*70)
print("\nMes | Consultantes | Consultas")
print("-" * 50)

for mes in sorted(months_data.keys()):
    data = months_data[mes]
    mes_nombre = meses_nombres.get(mes, f'Mes {mes}')
    print(f"{mes_nombre} | {len(data['consultantes_set'])} | {len(data['consultas'])}")
