#!/usr/bin/env python3
from openpyxl import load_workbook
from datetime import datetime

# Cargar el archivo
wb = load_workbook('agendamientos.xlsx')
ws = wb['2022']

print("ESTRUCTURA COMPLETA DE LA HOJA 2022\n")

# Ver todas las filas
for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=66, values_only=True), 1):
    # Si la fila tiene algún valor
    if any(row):
        # Mostrar la fila con índices de columna
        print(f"Fila {row_idx}:")
        for col_idx, cell in enumerate(row, 1):
            if cell:
                # Si es una fecha, mostrarla formateada
                if isinstance(cell, datetime):
                    print(f"  Col {col_idx}: {cell.strftime('%d/%m/%Y')} ({cell.strftime('%A')})")
                else:
                    print(f"  Col {col_idx}: {cell}")
        print()
